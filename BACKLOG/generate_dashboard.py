#!/usr/bin/env python3
"""
generate_dashboard.py
Reads BACKLOG/BACKLOG.xlsx and writes BACKLOG/dashboard.html — a read-only,
self-contained, auto-refreshing view of the backlog. No network required.

Run from the project root:  python3 BACKLOG/generate_dashboard.py
Claude Code should run this every time it changes the spreadsheet.
This script only READS the xlsx; it never writes to it.
"""

import json
import os
import sys
from datetime import datetime

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl is required:  pip3 install openpyxl --break-system-packages")

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "BACKLOG.xlsx")
OUT = os.path.join(HERE, "dashboard.html")

EXPECTED = ["ID", "Task", "Type", "Priority", "Status", "Source", "Sprint",
            "Component", "Details", "Depends On", "Date Added", "Date Completed",
            "Kick-Off Prompt", "Claude Notes"]


def load_rows():
    wb = load_workbook(XLSX, data_only=True)
    ws = wb["Backlog"] if "Backlog" in wb.sheetnames else wb.active
    values = list(ws.iter_rows(values_only=True))
    if not values:
        return []
    headers = [str(h).strip() if h is not None else "" for h in values[0]]
    rows = []
    for raw in values[1:]:
        if raw is None or all(c is None or str(c).strip() == "" for c in raw):
            continue
        row = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            val = raw[i] if i < len(raw) else None
            row[h] = "" if val is None else str(val).strip()
        if not row.get("ID") and not row.get("Task"):
            continue
        rows.append(row)
    return rows


def main():
    if not os.path.exists(XLSX):
        sys.exit(f"Spreadsheet not found: {XLSX}")
    rows = load_rows()
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    payload = {"rows": rows, "generatedAt": generated_at, "columns": EXPECTED}
    html = TEMPLATE.replace("/*__DATA__*/", json.dumps(payload))
    with open(OUT, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"Wrote {OUT}  ({len(rows)} item{'s' if len(rows) != 1 else ''}, generated {generated_at})")


TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Backlog</title>
<style>
  :root{
    --paper:#F4F5F7; --surface:#FFFFFF; --ink:#171B24; --muted:#6B7280;
    --faint:#9AA1AC; --hair:#E4E7EC; --navy:#2F5496; --navy-soft:#EAF0FA;
    --feature:#4338CA; --bug:#B91C1C; --refactor:#7C3AED; --chore:#475569;
    --hi:#C2410C; --med:#B45309; --lo:#64748B;
    --done:#15803D; --prog:#2F5496; --todo:#64748B; --blocked:#B45309;
    --claude:#0F766E; --me:#2F5496;
    --sans:-apple-system,BlinkMacSystemFont,"SF Pro Text","Segoe UI",Roboto,sans-serif;
    --mono:ui-monospace,"SF Mono",Menlo,Monaco,"Cascadia Code",monospace;
  }
  *{box-sizing:border-box}
  html,body{margin:0}
  body{background:var(--paper);color:var(--ink);font-family:var(--sans);
    font-size:14px;line-height:1.5;-webkit-font-smoothing:antialiased;}
  .wrap{max-width:1080px;margin:0 auto;padding:28px 24px 80px;}

  .top{display:flex;align-items:baseline;justify-content:space-between;
    gap:16px;flex-wrap:wrap;padding-bottom:16px;border-bottom:2px solid var(--navy);}
  .brand{font-family:var(--mono);font-weight:600;letter-spacing:.32em;
    font-size:15px;color:var(--navy);text-transform:uppercase;}
  .brand .dot{color:var(--faint);}
  .sync{display:flex;align-items:center;gap:14px;font-family:var(--mono);
    font-size:11.5px;color:var(--muted);}
  .pip{width:8px;height:8px;border-radius:50%;background:var(--done);
    display:inline-block;box-shadow:0 0 0 0 rgba(21,128,61,.5);}
  .pip.beat{animation:beat .9s ease-out;}
  @keyframes beat{0%{box-shadow:0 0 0 0 rgba(21,128,61,.5)}
    100%{box-shadow:0 0 0 7px rgba(21,128,61,0)}}
  .sync button{font:inherit;font-size:11.5px;color:var(--navy);background:none;
    border:1px solid var(--hair);border-radius:6px;padding:4px 9px;cursor:pointer;}
  .sync button:hover{background:var(--navy-soft);border-color:#C9D6EE;}
  .sync .toggle{color:var(--muted);cursor:pointer;}
  .sync .toggle.off{color:var(--faint);text-decoration:line-through;}

  .stats{display:flex;gap:10px;flex-wrap:wrap;margin:20px 0 8px;align-items:stretch;}
  .stat{background:var(--surface);border:1px solid var(--hair);border-radius:10px;
    padding:12px 16px;min-width:104px;}
  .stat .n{font-family:var(--mono);font-size:24px;font-weight:600;line-height:1;}
  .stat .l{font-size:11px;color:var(--muted);text-transform:uppercase;
    letter-spacing:.07em;margin-top:6px;}
  .stat.accent{background:#FFF7ED;border-color:#FED7AA;}
  .stat.accent .n{color:var(--hi);}
  .stat.src-me{background:var(--navy-soft);border-color:#D5E1F5;}
  .stat.src-me .n{color:var(--me);}
  .stat.src-claude{background:#E7F5F3;border-color:#BFE3DD;}
  .stat.src-claude .n{color:var(--claude);}
  .statgap{width:2px;background:var(--hair);margin:2px 4px;border-radius:2px;}

  .controls{margin:16px 0 4px;}
  .frow{display:flex;gap:6px;flex-wrap:wrap;align-items:center;margin-bottom:8px;}
  .flabel{font-size:11px;color:var(--faint);text-transform:uppercase;letter-spacing:.09em;
    font-weight:600;margin-right:4px;min-width:52px;}
  .chip{font:inherit;font-size:12px;color:var(--muted);background:var(--surface);
    border:1px solid var(--hair);border-radius:20px;padding:5px 12px;cursor:pointer;}
  .chip[aria-pressed="true"]{background:var(--navy);color:#fff;border-color:var(--navy);}
  .chip.src[aria-pressed="true"]{background:var(--claude);border-color:var(--claude);}

  .group{margin-top:26px;}
  .group h2{display:flex;align-items:center;gap:9px;font-size:12px;
    text-transform:uppercase;letter-spacing:.12em;color:var(--muted);
    margin:0 0 10px;font-weight:600;}
  .group h2 .count{font-family:var(--mono);color:var(--faint);letter-spacing:normal;}
  .group h2 .bar{height:1px;background:var(--hair);flex:1;}
  .group.blocked h2{color:var(--blocked);}

  .card{background:var(--surface);border:1px solid var(--hair);border-radius:12px;
    padding:14px 16px;margin-bottom:10px;display:grid;
    grid-template-columns:98px 1fr;gap:14px;}
  .rail{display:flex;flex-direction:column;gap:7px;align-items:flex-start;}
  .id{font-family:var(--mono);font-size:12px;font-weight:600;color:var(--navy);}
  .body h3{margin:0 0 3px;font-size:15px;font-weight:600;letter-spacing:-.01em;}
  .meta{font-size:12px;color:var(--muted);margin-bottom:8px;}
  .meta .comp{font-family:var(--mono);color:var(--faint);}
  .details{font-size:13px;color:#3B4250;margin:2px 0 0;}
  .tags{display:flex;gap:6px;flex-wrap:wrap;margin:8px 0 0;}
  .tag{font-family:var(--mono);font-size:10.5px;padding:2px 7px;border-radius:5px;
    background:#F1F3F7;color:#475569;border:1px solid var(--hair);}
  .tag.sprint{background:var(--navy-soft);color:var(--navy);border-color:#D5E1F5;}
  .tag.dep{background:#F1F3F7;color:#475569;}
  .tag.dep.wait{background:#FEF3E7;color:var(--blocked);border-color:#F6D9B0;}
  .tag.done-on{background:#EAF7EE;color:var(--done);border-color:#CDEBD6;}
  .tag.src-me{background:var(--navy-soft);color:var(--me);border-color:#D5E1F5;}
  .tag.src-claude{background:#E7F5F3;color:var(--claude);border-color:#BFE3DD;}

  .badge{display:inline-block;font-family:var(--mono);font-size:10.5px;
    font-weight:600;letter-spacing:.04em;text-transform:uppercase;
    padding:2px 7px;border-radius:5px;border:1px solid transparent;}
  .t-Feature{color:var(--feature);background:#EEF0FE;}
  .t-Bug{color:var(--bug);background:#FEECEC;}
  .t-Refactor{color:var(--refactor);background:#F3ECFE;}
  .t-Chore{color:var(--chore);background:#EEF1F5;}
  .p{font-family:var(--mono);font-size:11px;font-weight:600;}
  .p-High{color:var(--hi);} .p-Medium{color:var(--med);} .p-Low{color:var(--lo);}
  .p .tick{display:inline-block;width:6px;height:6px;border-radius:50%;
    margin-right:5px;vertical-align:middle;}
  .p-High .tick{background:var(--hi);} .p-Medium .tick{background:var(--med);}
  .p-Low .tick{background:var(--lo);}
  .st{font-family:var(--mono);font-size:10px;font-weight:600;letter-spacing:.03em;
    text-transform:uppercase;padding:2px 7px;border-radius:5px;}
  .st-NotStarted{color:var(--todo);background:#EEF1F5;}
  .st-InProgress{color:var(--prog);background:var(--navy-soft);}
  .st-Blocked{color:var(--blocked);background:#FEF3E7;}
  .st-Done{color:var(--done);background:#EAF7EE;}

  .note{margin-top:11px;border-left:3px solid var(--navy);background:var(--navy-soft);
    border-radius:0 8px 8px 0;padding:9px 12px;}
  .note .lbl{font-family:var(--mono);font-size:10px;letter-spacing:.1em;
    text-transform:uppercase;color:var(--navy);margin-bottom:4px;}
  .note .txt{font-size:13px;color:#233250;white-space:pre-wrap;}
  .note.empty{background:#FafBfC;border-left-color:var(--hair);}
  .note.empty .lbl{color:var(--faint);}
  .note.empty .txt{color:var(--faint);font-style:italic;}

  .empty-state{text-align:center;color:var(--muted);padding:60px 20px;
    font-family:var(--mono);font-size:13px;}
  .foot{margin-top:34px;font-family:var(--mono);font-size:11px;color:var(--faint);
    text-align:center;}

  @media (max-width:560px){
    .card{grid-template-columns:1fr;}
    .rail{flex-direction:row;flex-wrap:wrap;align-items:center;}
    .flabel{min-width:auto;width:100%;margin-bottom:2px;}
  }
  @media (prefers-reduced-motion:reduce){.pip.beat{animation:none;}}
</style>
</head>
<body>
<div class="wrap">
  <div class="top">
    <div class="brand">BACKLOG<span class="dot"> // </span>Zero Twenty One</div>
    <div class="sync">
      <span><span class="pip" id="pip"></span> synced <span id="stamp">&mdash;</span></span>
      <span class="toggle" id="auto">auto 45s</span>
      <button id="refresh">Refresh now</button>
    </div>
  </div>

  <div class="stats" id="stats"></div>
  <div class="controls" id="controls"></div>
  <div id="board"></div>

  <div class="foot" id="foot"></div>
</div>

<script id="payload" type="application/json">/*__DATA__*/</script>
<script>
(function(){
  var payload = JSON.parse(document.getElementById("payload").textContent);
  var rows = payload.rows || [];
  var esc = function(s){var d=document.createElement("div");d.textContent=s==null?"":s;return d.innerHTML;};

  // ---- state (persisted across auto-reloads) ----
  var STATE = {};
  try{ STATE = JSON.parse(sessionStorage.getItem("bl_state")||"{}"); }catch(e){}
  STATE.status = STATE.status || "All";
  STATE.type = STATE.type || "All";
  STATE.source = STATE.source || "All";
  STATE.sort = STATE.sort || "status";
  if(STATE.auto === undefined) STATE.auto = true;
  var saveState = function(){ try{sessionStorage.setItem("bl_state",JSON.stringify(STATE));}catch(e){} };

  var srcOf = function(r){ var s=(r.Source||"").trim(); return (s==="Me"||s==="Claude")?s:""; };

  // ---- stats ----
  var by = function(k,v){return rows.filter(function(r){return (r[k]||"")===v;}).length;};
  var openHigh = rows.filter(function(r){return (r.Status||"")!=="Done" && (r.Priority||"")==="High";}).length;
  var openMine = rows.filter(function(r){return srcOf(r)==="Me" && (r.Status||"")!=="Done";}).length;
  var openClaude = rows.filter(function(r){return srcOf(r)==="Claude" && (r.Status||"")!=="Done";}).length;
  function stat(n,l,cls){return '<div class="stat'+(cls?' '+cls:'')+'"><div class="n">'+n+'</div><div class="l">'+l+'</div></div>';}
  document.getElementById("stats").innerHTML =
    stat(by("Status","Not Started"),"Not started") +
    stat(by("Status","In Progress"),"In progress") +
    stat(by("Status","Blocked"),"Blocked") +
    stat(by("Status","Done"),"Done") +
    stat(openHigh,"Open \u00b7 high","accent") +
    '<div class="statgap"></div>' +
    stat(openMine,"Mine \u00b7 open","src-me") +
    stat(openClaude,"Claude \u00b7 open","src-claude");

  // ---- controls: status filter, type filter, source filter, sort ----
  var controls = document.getElementById("controls");
  function chipRow(label, dim, opts, isSrc){
    var row = document.createElement("div"); row.className="frow";
    if(label){ var l=document.createElement("span"); l.className="flabel"; l.textContent=label; row.appendChild(l); }
    opts.forEach(function(o){
      var b=document.createElement("button");
      b.className="chip"+(isSrc?" src":"");
      b.textContent=o.label; b.dataset.dim=dim; b.dataset.val=o.val;
      b.setAttribute("aria-pressed", STATE[dim]===o.val);
      b.onclick=function(){ STATE[dim]=o.val; saveState(); syncChips(); render(); };
      row.appendChild(b);
    });
    return row;
  }
  function opt(v,l){return {val:v,label:(l==null?v:l)};}
  controls.appendChild(chipRow("Status","status",
    [opt("All"),opt("Not Started"),opt("In Progress"),opt("Blocked"),opt("Done")]));
  controls.appendChild(chipRow("Type","type",
    [opt("All"),opt("Feature"),opt("Bug"),opt("Refactor"),opt("Chore")]));
  controls.appendChild(chipRow("Source","source",
    [opt("All","All sources"),opt("Me","Mine"),opt("Claude","Claude")], true));
  controls.appendChild(chipRow("Sort","sort",
    [opt("status","By status"),opt("newest","Newest first"),opt("oldest","Oldest first")]));
  function syncChips(){
    controls.querySelectorAll(".chip").forEach(function(c){
      c.setAttribute("aria-pressed", STATE[c.dataset.dim]===c.dataset.val);
    });
  }

  // ---- board ----
  var order = ["In Progress","Blocked","Not Started","Done"];
  var prioRank = {High:0,Medium:1,Low:2,"":3};
  function pr(p){ var v=prioRank[p]; return (v===undefined?3:v); }
  var statusOf = {};
  rows.forEach(function(r){ if(r.ID) statusOf[String(r.ID).trim()] = (r.Status||"").trim(); });
  function parseDeps(s){ return String(s||"").split(/[,\s]+/).map(function(x){return x.trim();}).filter(Boolean); }
  function dateVal(r){ var t=Date.parse(r["Date Added"]||""); return isNaN(t)?0:t; }

  function passFilters(r){
    var okS = STATE.status==="All" || (r.Status||"")===STATE.status;
    var okT = STATE.type==="All" || (r.Type||"")===STATE.type;
    var okSrc = STATE.source==="All" || srcOf(r)===STATE.source;
    return okS && okT && okSrc;
  }

  function render(){
    var board=document.getElementById("board"); board.innerHTML="";
    var pool = rows.filter(passFilters);
    if(!pool.length){ board.innerHTML='<div class="empty-state">Nothing matches these filters.</div>'; return; }

    if(STATE.sort==="status"){
      order.forEach(function(st){
        if(STATE.status!=="All" && STATE.status!==st) return;
        var items = pool.filter(function(r){return (r.Status||"")===st;});
        if(!items.length) return;
        items.sort(function(a,b){return pr(a.Priority)-pr(b.Priority);});
        var g=document.createElement("div"); g.className="group"+(st==="Blocked"?" blocked":"");
        g.innerHTML='<h2>'+esc(st)+' <span class="count">'+items.length+'</span><span class="bar"></span></h2>';
        items.forEach(function(r){ g.appendChild(card(r,false)); });
        board.appendChild(g);
      });
    } else {
      var newest = STATE.sort==="newest";
      var items = pool.slice().sort(function(a,b){ return newest ? dateVal(b)-dateVal(a) : dateVal(a)-dateVal(b); });
      var g=document.createElement("div"); g.className="group";
      g.innerHTML='<h2>'+(newest?"Newest first":"Oldest first")+' <span class="count">'+items.length+'</span><span class="bar"></span></h2>';
      items.forEach(function(r){ g.appendChild(card(r,true)); });
      board.appendChild(g);
    }
  }

  function card(r, showStatus){
    var el=document.createElement("div"); el.className="card"; el.dataset.id=r.ID||"";
    var type=r.Type||""; var prio=r.Priority||""; var st=(r.Status||"");
    var note=(r["Claude Notes"]||"").trim();
    var stClass = st.replace(/\s+/g,"");
    var rail='<div class="rail">'+
      '<span class="id">'+esc(r.ID||"\u2014")+'</span>'+
      (type?'<span class="badge t-'+esc(type)+'">'+esc(type)+'</span>':'')+
      (prio?'<span class="p p-'+esc(prio)+'"><span class="tick"></span>'+esc(prio)+'</span>':'')+
      (showStatus && st?'<span class="st st-'+esc(stClass)+'">'+esc(st)+'</span>':'')+
      '</div>';
    var comp=r.Component?'<span class="comp">'+esc(r.Component)+'</span> \u00b7 ':'';
    var added=r["Date Added"]?'added '+esc(r["Date Added"]):'';
    var meta='<div class="meta">'+comp+added+'</div>';
    var details=r.Details?'<p class="details">'+esc(r.Details)+'</p>':'';

    var tags=[];
    var src=srcOf(r);
    if(src) tags.push('<span class="tag '+(src==="Claude"?"src-claude":"src-me")+'">'+(src==="Claude"?"Claude":"You")+'</span>');
    var sprint=(r.Sprint||"").trim();
    if(sprint && sprint.toLowerCase()!=="unscheduled") tags.push('<span class="tag sprint">'+esc(sprint)+'</span>');
    parseDeps(r["Depends On"]).forEach(function(dep){
      var depDone=(statusOf[dep]||"")==="Done";
      tags.push('<span class="tag dep'+(depDone?"":" wait")+'">'+(depDone?"\u2713 ":"\u29d6 ")+'needs '+esc(dep)+'</span>');
    });
    var done=(r["Date Completed"]||"").trim();
    if(done) tags.push('<span class="tag done-on">\u2713 completed '+esc(done)+'</span>');
    var tagsHtml=tags.length?'<div class="tags">'+tags.join("")+'</div>':'';

    var noteHtml=note
      ? '<div class="note"><div class="lbl">Claude notes</div><div class="txt">'+esc(note)+'</div></div>'
      : '<div class="note empty"><div class="lbl">Claude notes</div><div class="txt">No notes yet</div></div>';
    el.innerHTML=rail+'<div class="body"><h3>'+esc(r.Task||"(untitled)")+'</h3>'+meta+details+tagsHtml+noteHtml+'</div>';
    return el;
  }

  // ---- sync line + auto refresh ----
  document.getElementById("stamp").textContent = payload.generatedAt || "\u2014";
  document.getElementById("pip").classList.add("beat");
  document.getElementById("foot").textContent =
    rows.length+" item"+(rows.length!==1?"s":"")+" \u00b7 generated "+(payload.generatedAt||"")+" \u00b7 read-only view";

  var autoEl=document.getElementById("auto");
  function paintAuto(){ autoEl.classList.toggle("off",!STATE.auto); autoEl.textContent=STATE.auto?"auto 45s":"auto off"; }
  autoEl.onclick=function(){ STATE.auto=!STATE.auto; saveState(); paintAuto(); schedule(); };
  paintAuto();

  function doReload(){ try{sessionStorage.setItem("bl_scroll",String(window.scrollY));}catch(e){} location.reload(); }
  document.getElementById("refresh").onclick=doReload;
  var timer=null;
  function schedule(){ if(timer){clearTimeout(timer);timer=null;} if(STATE.auto){ timer=setTimeout(doReload,45000); } }

  try{ var y=sessionStorage.getItem("bl_scroll"); if(y!==null){ window.scrollTo(0,parseInt(y,10)||0); sessionStorage.removeItem("bl_scroll"); } }catch(e){}

  syncChips(); render(); schedule();
})();
</script>
</body>
</html>
"""

if __name__ == "__main__":
    main()
